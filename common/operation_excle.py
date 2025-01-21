from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font  # 导入字体模块
import openpyxl, string
from common.logger import Logger


class operation_excle:

    @classmethod
    def read_excel(cls, file_name, case_severity_list):

        wb = openpyxl.load_workbook(file_name)
        # ws = wb.active#打开当前页
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        rows_list = []
        for title in sheet_names:
            ws = wb[title]  # 打开指定页
            # 取出每行的值，以list方式存放

            # rows_list=[[],[],[]]
            for row in ws.rows:
                row_list = []
                for cell in row:
                    if cell.value == None:
                        cell_srt = ''
                    else:
                        cell_srt = cell.value
                    # print(cell.value)
                    row_list.append(cell_srt)

                # row_list.append(title)
                row_list.insert(0, title)

                if row_list[5] in case_severity_list:  # 筛选匹配的用例等级进行测试

                    rows_list.append(row_list)
            # print(rows_list)
            # 结果转换成键值对的形式存放

        result = []
        for i in range(len(rows_list) - 1):
            row_dict = {}
            for j in range(len(rows_list[0])):
                row_dict[rows_list[0][j]] = rows_list[i + 1][j]
            result.append(row_dict)

        # return result#返回字典形式

        return rows_list  # 返回列表形式

    def str_count_to_width(self, str):
        '''找出字符串中的中英文、空格、数字、标点符号个数'''
        count_en = count_dg = count_sp = count_zh = count_pu = 0
        for s in str:
            # 英文
            if s in string.ascii_letters:
                count_en += 1
            # 数字
            elif s.isdigit():
                count_dg += 1
            # 空格
            elif s.isspace():
                count_sp += 1
            # 中文，除了英文之外，剩下的字符认为就是中文
            elif s.isalpha():
                count_zh += 1
            # 特殊字符
            else:
                count_pu += 1

        width = count_en * 1 + count_dg * 1.5 + count_sp * 1 + count_zh * 2 + count_pu * 1.25  # 宋体 11号的参数，其他格式可自己尝试
        return (float('%.2f' % width))
